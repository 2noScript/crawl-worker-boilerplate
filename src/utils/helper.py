from openpyxl import Workbook,load_workbook
import json
import re
import aiohttp
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pathlib import Path

def read_file_lines(file_path):
    path = Path(file_path)
    data = []
    with path.open('r') as file:
        for line in file:
            data.append(line.strip())
    return data



def get_sheet_id_from_url(url):
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if match:
        return match.group(1)
    return None



async def get_url_links(sheet_url):

    sheet_id = get_sheet_id_from_url(sheet_url)
    if not sheet_id:
        return []
    
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:json"
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                data = await response.text()
        
        match = re.search(r"google.visualization.Query.setResponse\((.*)\);", data)
        if match:
            json_data = json.loads(match.group(1))
            rows = json_data["table"]["rows"]
            
            first_column_values = []
            for row in rows:
                first_cell = row["c"][0]
                first_column_value = first_cell.get("v", "") if first_cell is not None else ""
                # Only add the value if it's a URL (starts with http:// or https://)
                if isinstance(first_column_value, str) and (first_column_value.startswith("http://") or first_column_value.startswith("https://")):
                    first_column_values.append(first_column_value)
            
            return first_column_values
    except Exception as e:
        print(f"Error fetching sheet data: {e}")
    
    return []




def export_to_excel(filename, sheet_name, headers, data):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(headers)

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")

    for cell in ws[1]: 
        cell.font = bold_font
        cell.alignment = center_alignment

    for row in data:
        ws.append(row)
    wb.save(filename)

